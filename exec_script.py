import xml.etree.cElementTree as ET
from openpyxl import load_workbook
from tools import file_operations
from config import reader_config as cfg


def xl_to_xml_for_testlink(columns_in_use: dict, rows_to_skip: int, folder_xl_file_list: str) -> None:
    for file_path in folder_xl_file_list:
        workbook = load_workbook(file_path)
        iter_rows = workbook.active.iter_rows(min_row=rows_to_skip)

        main_test_suite = None
        number_of_current_test_suite = 1

        for row in iter_rows:
            for cell in row:
                if cell.value is not None and cell.column_letter in columns_in_use:
                    column_role = columns_in_use[cell.column_letter]["column_role"]

                    if column_role == 'main_test_suite':
                        if main_test_suite is not None:
                            file_name = file_operations.prepare_file_name(
                                file_path, number_of_current_test_suite)
                            file_operations.write_tree_to_xml_file(
                                main_test_suite, file_name)
                            number_of_current_test_suite += 1
                        main_test_suite = ET.Element('testsuite', name=cell.value)
                        parent_of_test_case = main_test_suite
                    elif column_role == 'sub_test_suite':
                        parent_of_test_case = ET.SubElement(main_test_suite, 'testsuite', name=cell.value)
                    elif column_role == 'test_case':
                        test_case = ET.SubElement(parent_of_test_case, 'testcase', name=cell.value)
                    elif column_role in ['importance', 'summary', 'preconditions']:
                        ET.SubElement(test_case, column_role).text = str(cell.value)
                        if column_role == 'preconditions':
                            steps = ET.SubElement(test_case, 'steps')
                            case_step = 1
                    elif column_role == 'actions':
                        step = ET.SubElement(steps, 'step')
                        ET.SubElement(step, 'step_number').text = str(case_step)
                        ET.SubElement(step, 'actions').text = str(cell.value)
                    elif column_role == 'expected_results':
                        ET.SubElement(step, 'expectedresults').text = str(cell.value)
                        case_step += 1

        file_name = file_operations.prepare_file_name(
            file_path, number_of_current_test_suite)
        file_operations.write_tree_to_xml_file(
            main_test_suite, file_name)
        number_of_current_test_suite += 1


if __name__ == "__main__":
    folder_xl_file_list = file_operations.list_files(cfg.file_format)
    xl_to_xml_for_testlink(
        cfg.columns_in_use, cfg.rows_to_skip, folder_xl_file_list)
